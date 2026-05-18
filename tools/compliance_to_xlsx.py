#!/usr/bin/env python3
"""
tools/compliance_to_xlsx.py — Convert working/compliance-matrix.md to Excel.

Produces final/xlsx/compliance-matrix.xlsx with:
  - Sheet 1 "Matrix"   — all rows, filter dropdowns, color-coded Status column
  - Sheet 2 "Summary"  — counts by status group
  - Sheet 3 "Gaps"     — rows with status Gap, Partial, or Exception

Usage:
    python tools/compliance_to_xlsx.py --proposal nato-diana-decision-superiority
    python tools/compliance_to_xlsx.py --proposal myproposal --workspace /path/to/workspace
"""

import argparse
import re
import sys
import io
from pathlib import Path

# Force UTF-8 on Windows consoles that default to cp1252
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.filters import AutoFilter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


WORKSPACE_ROOT = Path(__file__).parent.parent

# ── Status colour palette ─────────────────────────────────────────────────────
STATUS_FILLS = {
    "Covered":   PatternFill("solid", fgColor="C6EFCE"),   # green
    "Drafted":   PatternFill("solid", fgColor="BDD7EE"),   # blue
    "Planned":   PatternFill("solid", fgColor="FFFACD"),   # light yellow
    "Partial":   PatternFill("solid", fgColor="FFDAB9"),   # orange/peach
    "Exception": PatternFill("solid", fgColor="E0E0E0"),   # grey
    "Gap":       PatternFill("solid", fgColor="FFC7CE"),   # red
}
STATUS_FONTS = {
    "Covered":   Font(bold=False, color="276221"),
    "Drafted":   Font(bold=False, color="1F4E79"),
    "Planned":   Font(bold=False, color="7D6608"),
    "Partial":   Font(bold=False, color="7D4400"),
    "Exception": Font(bold=False, color="444444"),
    "Gap":       Font(bold=False, color="9C0006"),
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(size=10)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ── Markdown table parser ─────────────────────────────────────────────────────

def _is_separator(line: str) -> bool:
    return bool(re.match(r'^\s*\|[\s\-:|]+\|\s*$', line))


def parse_compliance_matrix(md_text: str) -> tuple[list[str], list[list[str]]]:
    """
    Extract header row and data rows from the compliance matrix markdown.
    Skips separator rows, comment lines, and non-table lines.
    Returns (headers, rows) — both as lists of strings.
    """
    # Strip HTML comments
    md_text = re.sub(r'<!--.*?-->', '', md_text, flags=re.DOTALL)

    headers: list[str] = []
    rows: list[list[str]] = []
    in_table = False

    for line in md_text.split('\n'):
        s = line.strip()
        if not s.startswith('|'):
            in_table = False
            continue
        if _is_separator(s):
            in_table = True  # separator follows the header row
            continue
        cells = [c.strip() for c in s.strip('|').split('|')]
        if not in_table:
            # First pipe row before any separator = header
            headers = cells
        else:
            rows.append(cells)

    return headers, rows


# ── Worksheet helpers ─────────────────────────────────────────────────────────

def _apply_header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30


def _write_data_rows(ws, headers: list[str], rows: list[list[str]],
                     status_col_idx: int | None) -> None:
    for row_data in rows:
        # Pad or trim to match header count
        padded = (row_data + [""] * len(headers))[:len(headers)]
        ws.append(padded)
        row_idx = ws.max_row
        for c_idx, cell in enumerate(ws[row_idx]):
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER
            # Status colouring
            if status_col_idx is not None and c_idx == status_col_idx:
                status_val = (padded[c_idx] if c_idx < len(padded) else "").strip()
                fill = STATUS_FILLS.get(status_val)
                font = STATUS_FONTS.get(status_val)
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font


def _set_column_widths(ws, headers: list[str]) -> None:
    """Heuristic column widths based on header name."""
    widths = {
        "req id":      10,
        "source":      14,
        "requirement": 50,
        "section":     22,
        "page":         8,
        "status":      12,
        "evidence":    35,
        "notes":       30,
        "coverage":    12,
        "gap":         20,
    }
    for c_idx, hdr in enumerate(headers):
        key = hdr.lower()
        width = 15  # default
        for k, w in widths.items():
            if k in key:
                width = w
                break
        ws.column_dimensions[get_column_letter(c_idx + 1)].width = width


def _find_status_col(headers: list[str]) -> int | None:
    for i, h in enumerate(headers):
        if "status" in h.lower():
            return i
    return None


def _add_autofilter(ws, headers: list[str]) -> None:
    if not headers:
        return
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"


# ── Summary sheet ─────────────────────────────────────────────────────────────

def _build_summary(ws_sum, rows: list[list[str]], headers: list[str],
                   status_col_idx: int | None) -> None:
    ws_sum.column_dimensions["A"].width = 18
    ws_sum.column_dimensions["B"].width = 12

    ws_sum.append(["Status", "Count"])
    for cell in ws_sum[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    counts: dict[str, int] = {}
    if status_col_idx is not None:
        for row in rows:
            val = row[status_col_idx].strip() if status_col_idx < len(row) else ""
            counts[val] = counts.get(val, 0) + 1

    order = ["Covered", "Drafted", "Planned", "Partial", "Exception", "Gap"]
    for status in order:
        n = counts.get(status, 0)
        ws_sum.append([status, n])
        row_idx = ws_sum.max_row
        fill = STATUS_FILLS.get(status)
        font = STATUS_FONTS.get(status)
        for cell in ws_sum[row_idx]:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            cell.border = THIN_BORDER

    # Other statuses not in the canonical list
    for status, n in sorted(counts.items()):
        if status not in order:
            ws_sum.append([status or "(blank)", n])

    # Total
    ws_sum.append(["TOTAL", len(rows)])
    r = ws_sum.max_row
    ws_sum[f"A{r}"].font = Font(bold=True, size=10)
    ws_sum[f"B{r}"].font = Font(bold=True, size=10)
    for cell in ws_sum[r]:
        cell.border = THIN_BORDER


# ── Gaps sheet ────────────────────────────────────────────────────────────────

def _build_gaps(ws_gaps, headers: list[str], rows: list[list[str]],
                status_col_idx: int | None) -> None:
    gap_statuses = {"Gap", "Partial", "Exception"}
    gap_rows = []
    if status_col_idx is not None:
        gap_rows = [
            r for r in rows
            if (r[status_col_idx].strip() if status_col_idx < len(r) else "") in gap_statuses
        ]

    _apply_header_row(ws_gaps, headers)
    _write_data_rows(ws_gaps, headers, gap_rows, status_col_idx)
    _set_column_widths(ws_gaps, headers)

    if not gap_rows:
        ws_gaps.append(["(No gaps, partials, or exceptions — clean compliance!)"])


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Convert working/compliance-matrix.md to Excel (.xlsx).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--proposal", required=True,
                        help="Proposal slug (directory under proposals/)")
    parser.add_argument("--workspace", default=None,
                        help="Workspace root path (default: parent of tools/)")
    parser.add_argument("--input", default=None,
                        help="Explicit path to compliance-matrix.md (overrides default)")
    args = parser.parse_args()

    workspace = Path(args.workspace) if args.workspace else WORKSPACE_ROOT
    proposal_dir = workspace / "proposals" / args.proposal

    if not proposal_dir.exists():
        print(f"ERROR: Proposal directory not found: {proposal_dir}", file=sys.stderr)
        sys.exit(1)

    md_path = Path(args.input) if args.input else proposal_dir / "working" / "compliance-matrix.md"
    if not md_path.exists():
        print(f"ERROR: Compliance matrix not found: {md_path}", file=sys.stderr)
        sys.exit(1)

    out_dir = proposal_dir / "final" / "xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "compliance-matrix.xlsx"

    # Parse
    md_text = md_path.read_text(encoding="utf-8")
    headers, rows = parse_compliance_matrix(md_text)

    if not headers:
        print("ERROR: No table found in compliance-matrix.md. "
              "Ensure the file contains a pipe-table with a header row.", file=sys.stderr)
        sys.exit(1)

    status_col_idx = _find_status_col(headers)

    print(f"Proposal : {args.proposal}")
    print(f"Input    : {md_path}")
    print(f"Columns  : {headers}")
    print(f"Rows     : {len(rows)}")
    print(f"Status col: {headers[status_col_idx] if status_col_idx is not None else '(not found)'}")
    print()

    # Build workbook
    wb = openpyxl.Workbook()

    # Sheet 1 — Matrix
    ws_matrix = wb.active
    ws_matrix.title = "Matrix"
    ws_matrix.freeze_panes = "A2"
    _apply_header_row(ws_matrix, headers)
    _write_data_rows(ws_matrix, headers, rows, status_col_idx)
    _set_column_widths(ws_matrix, headers)
    _add_autofilter(ws_matrix, headers)

    # Sheet 2 — Summary
    ws_sum = wb.create_sheet("Summary")
    _build_summary(ws_sum, rows, headers, status_col_idx)

    # Sheet 3 — Gaps
    ws_gaps = wb.create_sheet("Gaps")
    _build_gaps(ws_gaps, headers, rows, status_col_idx)

    wb.save(out_path)
    print(f"[OK] {out_path}")
    print(f"     Matrix sheet: {len(rows)} rows")

    if status_col_idx is not None:
        from collections import Counter
        counts = Counter(
            (r[status_col_idx].strip() if status_col_idx < len(r) else "") for r in rows
        )
        for k in ["Covered", "Drafted", "Planned", "Partial", "Exception", "Gap"]:
            if counts.get(k, 0):
                print(f"     {k:12s}: {counts[k]}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
