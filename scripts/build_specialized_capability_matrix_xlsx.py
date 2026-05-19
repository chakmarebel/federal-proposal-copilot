"""Build a formatted engineering-review workbook from a capability matrix."""
import argparse

import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
NAVY       = "1F3864"
LIGHT_BLUE = "DCE6F1"
WHITE      = "FFFFFF"
GREEN      = "C6EFCE"
YELLOW     = "FFEB9C"
RED        = "FFC7CE"

HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=10)
BODY_FONT   = Font(name="Arial", size=10)
BOLD_FONT   = Font(name="Arial", bold=True, size=10)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
ALT_FILL    = PatternFill("solid", fgColor=LIGHT_BLUE)
WHITE_FILL  = PatternFill("solid", fgColor=WHITE)
GREEN_FILL  = PatternFill("solid", fgColor=GREEN)
YELLOW_FILL = PatternFill("solid", fgColor=YELLOW)
RED_FILL    = PatternFill("solid", fgColor=RED)

WRAP   = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

# ---------------------------------------------------------------------------
# Coverage / severity colouring
# ---------------------------------------------------------------------------
COVERAGE_MAP = {
    "full":        GREEN_FILL,
    "full + 🏆":  GREEN_FILL,
    "✅ full":     GREEN_FILL,
    "✅ full + 🏆": GREEN_FILL,
    "partial":     YELLOW_FILL,
    "🟡 partial":  YELLOW_FILL,
    "gap":         RED_FILL,
    "🔴 gap":      RED_FILL,
    "medium":      YELLOW_FILL,
    "tight":       YELLOW_FILL,
    "🟡 tight":    YELLOW_FILL,
    "achievable":  GREEN_FILL,
    "🟢 achievable": GREEN_FILL,
    "high":        RED_FILL,
    "low-medium":  YELLOW_FILL,
    "low":         GREEN_FILL,
    "competitive": GREEN_FILL,
}

def coverage_fill(value: str):
    key = value.strip().lower()
    for k, fill in COVERAGE_MAP.items():
        if k in key:
            return fill
    return None

def strip_emoji(text: str) -> str:
    """Remove leading emoji + coverage badge prefixes for clean cell values."""
    text = re.sub(r"^[✅🟡🔴🏆🟢]\s*", "", text.strip())
    return text

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def apply_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = thin_border()
    ws.row_dimensions[row].height = 28

def apply_body_row(ws, row_idx, values, alt=False, coverage_cols=None):
    base_fill = ALT_FILL if alt else WHITE_FILL
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=col, value=val)
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = thin_border()
        if coverage_cols and col in coverage_cols:
            cf = coverage_fill(val or "")
            c.fill = cf if cf else base_fill
        else:
            c.fill = base_fill

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def auto_width(ws, caps: dict = None):
    caps = caps or {}
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        header = str(col_cells[0].value or "")
        cap = caps.get(header, 40)
        max_len = max((len(line) for cell in col_cells if cell.value
                       for line in str(cell.value).split("\n")), default=8)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), cap)

# ---------------------------------------------------------------------------
# Markdown table parser
# ---------------------------------------------------------------------------
def parse_md_table(lines):
    table_lines = [l for l in lines if l.strip().startswith("|")]
    if not table_lines:
        return [], []
    headers = [c.strip() for c in table_lines[0].strip("|").split("|")]
    rows = []
    for line in table_lines[2:]:
        cells = [c.strip() for c in line.strip("|").split("|")]
        while len(cells) < len(headers):
            cells.append("")
        rows.append(cells[:len(headers)])
    return headers, rows

# ---------------------------------------------------------------------------
# Section splitter — matches "## N. Title" headings
# ---------------------------------------------------------------------------
def split_sections(text):
    sections = {}
    current_key = None
    current_title = ""
    current_lines = []
    sec_re = re.compile(r"^##\s+(\d+)\.\s+(.*)")
    for line in text.splitlines():
        m = sec_re.match(line)
        if m:
            if current_key is not None:
                sections[int(current_key)] = (current_title, current_lines)
            current_key = m.group(1)
            current_title = m.group(2).strip()
            current_lines = []
        elif current_key is not None:
            current_lines.append(line)
    if current_key is not None:
        sections[int(current_key)] = (current_title, current_lines)
    return sections

# ---------------------------------------------------------------------------
# Clean rows: strip emoji from Coverage column values
# ---------------------------------------------------------------------------
def clean_rows(headers, rows, coverage_col_names=("Coverage", "Feasibility", "Coverage Level", "Severity")):
    cov_indices = {i for i, h in enumerate(headers) if any(c in h for c in coverage_col_names)}
    cleaned = []
    for row in rows:
        new_row = list(row)
        for i in cov_indices:
            if i < len(new_row):
                new_row[i] = strip_emoji(new_row[i])
        cleaned.append(new_row)
    return cleaned

# ---------------------------------------------------------------------------
# Build each sheet
# ---------------------------------------------------------------------------
def add_table_sheet(wb, title, headers, rows, coverage_col_names=("Coverage", "Feasibility", "Severity")):
    ws = wb.create_sheet(title=title)
    ws.freeze_panes = "A2"
    apply_header(ws, headers)
    rows = clean_rows(headers, rows, coverage_col_names)
    cov_cols = {i + 1 for i, h in enumerate(headers) if any(c in h for c in coverage_col_names)}
    for i, row in enumerate(rows):
        apply_body_row(ws, i + 2, row, alt=(i % 2 == 1), coverage_cols=cov_cols)
    auto_width(ws)
    return ws

def add_summary_sheet(wb, meta, sections):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.freeze_panes = "A3"

    # Title banner
    title_cell = ws.cell(row=1, column=1, value="CAPABILITY MATRIX — DIGITAL GUARDIAN SIMULATOR")
    title_cell.font = Font(name="Arial", bold=True, color=WHITE, size=13)
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 36

    # Subtitle
    sub = ws.cell(row=2, column=1, value="Engineering Review Package — [Your Company], Inc. vs. SOW Requirements")
    sub.font = Font(name="Arial", italic=True, size=10)
    sub.fill = ALT_FILL
    sub.alignment = CENTER
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 22

    fields = [
        ("Proposal Slug",       meta.get("slug", "example-proposal")),
        ("Customer",            "U.S. Space Force / Defense Acquisition University (DAU)"),
        ("Contract Vehicle",    "Example Marketplace OTA (Status: Awardable)"),
        ("Funding Range",       "$125k–$200k NTE (FFP prototype)"),
        ("Demo Target",         "August 2026 — NVIDIA DGX Spark (100% local)"),
        ("Generated",           meta.get("generated", "2026-05-14")),
        ("Overall Coverage",    meta.get("coverage", "96% (24/25 requirements)")),
        ("Net-New Dev Estimate","~6.5 weeks of engineering effort (parallelized across 2–3 engineers)"),
    ]
    row = 3
    for i, (label, value) in enumerate(fields):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = BOLD_FONT
        lc.fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        lc.border = thin_border()
        lc.alignment = WRAP
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = BODY_FONT
        vc.fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        vc.border = thin_border()
        vc.alignment = WRAP
        ws.merge_cells(f"B{row}:E{row}")
        row += 1

    # Scorecard sub-table
    row += 1
    sc_hdr = ws.cell(row=row, column=1, value="Coverage Scorecard")
    sc_hdr.font = HEADER_FONT; sc_hdr.fill = HEADER_FILL; sc_hdr.alignment = CENTER; sc_hdr.border = thin_border()
    ws.merge_cells(f"A{row}:E{row}")
    ws.row_dimensions[row].height = 24
    row += 1

    scorecard_headers = ["Category", "Full ✅", "Partial 🟡", "Gap 🔴", "Total"]
    scorecard_rows = [
        ["Technical / Deployment", "6", "3", "0", "9"],
        ["Functional / Application", "5", "5", "1", "11"],
        ["Deliverables", "3", "2", "0", "5"],
        ["TOTAL", "14", "10", "1", "25"],
    ]
    for col, h in enumerate(scorecard_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = thin_border()
    row += 1
    for i, sr in enumerate(scorecard_rows):
        for col, val in enumerate(sr, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = BOLD_FONT if sr[0] == "TOTAL" else BODY_FONT
            c.fill = ALT_FILL if i % 2 == 1 else WHITE_FILL
            c.alignment = CENTER
            c.border = thin_border()
        row += 1

    # Gap summary header
    row += 1
    gh = ws.cell(row=row, column=1, value="Gap Summary (see '6-Gap Summary' sheet for full detail)")
    gh.font = HEADER_FONT; gh.fill = HEADER_FILL; gh.alignment = CENTER; gh.border = thin_border()
    ws.merge_cells(f"A{row}:E{row}")
    ws.row_dimensions[row].height = 24
    row += 1

    gap_headers = ["Gap", "Severity", "Est. Build Effort", "Mitigation"]
    gap_rows = [
        ["Neo4j graph DB integration", "Medium", "~1 week", "Start Week 1 parallel to persona work; use Neo4j Docker image"],
        ["GO/SES acquisition persona tuning", "Medium", "~1 week", "Leverage Acquisitions LoRA as base; 4 persona profiles via prompt engineering"],
        ["War game orchestration state machine", "Medium", "~1.5 weeks", "Build on existing agentic framework; define state machine in Week 1"],
        ["Why visualization UI (dynamic bars)", "Medium — new build", "~1 week", "React/HTML/JS; independent of LLM work — parallelize with persona work"],
        ["Injection API + WebSocket real-time push", "Low-Medium", "~1 week", "Dependent on visualization UI completion"],
        ["Zero Trust / audit logging spec", "Low", "~0.5 weeks", "Architecture section in proposal; implementation Week 6–7"],
        ["DAU SME data (GFI — external dependency)", "High (gov-owned)", "Gov furnished", "Define GFD deadline of Week 1 in proposal; accept partial graph at demo if delayed"],
    ]
    for col, h in enumerate(gap_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = thin_border()
    ws.merge_cells(f"D{row}:E{row}")
    row += 1
    for i, gr in enumerate(gap_rows):
        sev = gr[1]
        sev_fill = coverage_fill(sev) or (ALT_FILL if i % 2 == 1 else WHITE_FILL)
        for col, val in enumerate(gr, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = BODY_FONT
            c.fill = sev_fill if col == 2 else (ALT_FILL if i % 2 == 1 else WHITE_FILL)
            c.alignment = WRAP
            c.border = thin_border()
        ws.merge_cells(f"D{row}:E{row}")
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 22
    return ws

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Build a specialized capability-matrix workbook from markdown."
    )
    parser.add_argument("input", type=Path, help="Path to working/capability-matrix.md")
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        help="Output .xlsx path. Defaults to final/xlsx/capability-matrix.xlsx next to the proposal.",
    )
    args = parser.parse_args()

    md_path = args.input
    proposal_root = md_path.parent.parent if md_path.parent.name == "working" else md_path.parent
    out_path = args.output or proposal_root / "final" / "xlsx" / "capability-matrix.xlsx"

    text = md_path.read_text(encoding="utf-8")
    sections = split_sections(text)

    meta = {
        "slug":     proposal_root.name,
        "generated":"2026-05-14",
        "coverage": "96% (24/25 requirements; 1 gap = Why visualization UI)",
    }

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    add_summary_sheet(wb, meta, sections)

    sheet_config = [
        (1, "1-Tech & Deployment",    ("Coverage",)),
        (2, "2-Functional Reqs",      ("Coverage",)),
        (3, "3-Deliverables",         ("Coverage",)),
        (4, "4-Milestone Feasibility",("Feasibility",)),
        (5, "5-Competitive Summary",  ()),
        (6, "6-Gap Summary",          ("Severity",)),
        (7, "7-Coverage Scorecard",   ("Full ✅", "Partial 🟡", "Gap 🔴")),
    ]

    for sec_num, sheet_name, cov_names in sheet_config:
        _, lines = sections.get(sec_num, ("", []))
        headers, rows = parse_md_table(lines)
        if headers:
            add_table_sheet(wb, sheet_name, headers, rows, coverage_col_names=cov_names)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"Saved: {out_path}")

if __name__ == "__main__":
    main()
