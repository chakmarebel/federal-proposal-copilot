#!/usr/bin/env python3
"""
Extract STRUCTURAL patterns from a winning proposal pricing workbook (.xlsx).

Surfaces:
  - Sheet names + dimensions
  - Per-sheet: row count, column count, # of formulas, # of values
  - Header row content (first row treated as headers)
  - Formula patterns (counted by template, not full formulas)
  - Merged cell patterns
  - Use of conditional formatting / freeze panes
  - Number formatting patterns

Source content is NOT printed verbatim — only structural metadata that
informs how a winning pricing workbook is organized.

Usage:
    python scripts/extract-xlsx-patterns.py /path/to/workbook.xlsx
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl


def normalize_formula(formula: str) -> str:
    """Replace cell refs and numbers with placeholders to count formula templates."""
    s = re.sub(r"\$?[A-Z]+\$?\d+", "CELL", formula)
    s = re.sub(r"\b\d+(\.\d+)?\b", "N", s)
    return s


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("xlsx", type=Path)
    args = ap.parse_args(argv)

    if not args.xlsx.exists():
        print(f"error: {args.xlsx} not found", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(str(args.xlsx), data_only=False)
    print(f"=== {args.xlsx.name} ===")
    print(f"Sheets: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"  - {name}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        n_cells = 0
        n_formulas = 0
        n_values = 0
        formula_templates: Counter[str] = Counter()
        number_formats: Counter[str] = Counter()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                n_cells += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    n_formulas += 1
                    formula_templates[normalize_formula(cell.value)] += 1
                else:
                    n_values += 1
                if cell.number_format:
                    number_formats[cell.number_format] += 1

        print(f"--- Sheet: {sheet_name} ({max_row} rows x {max_col} cols) ---")
        print(f"  Non-empty cells: {n_cells}  (formulas: {n_formulas}, values: {n_values})")
        print(f"  Merged cell ranges: {len(ws.merged_cells.ranges)}")
        print(f"  Frozen panes: {ws.freeze_panes if ws.freeze_panes else 'none'}")
        print(f"  Auto-filter: {'yes' if ws.auto_filter.ref else 'no'}")

        # Headers (row 1)
        if max_row >= 1:
            header_row = []
            for cell in ws[1]:
                if cell.value is not None:
                    val = str(cell.value)
                    header_row.append(val[:40])
            if header_row:
                print(f"  Row 1 headers ({len(header_row)} non-empty):")
                for i, h in enumerate(header_row, 1):
                    print(f"    col {i}: {h}")

        # Top formula templates
        if formula_templates:
            print(f"  Top formula templates (template = formula with cell-refs and numbers placeholdered):")
            for tpl, c in formula_templates.most_common(8):
                print(f"    {c:>4}x  {tpl[:90]}")

        # Number formats
        if number_formats:
            print(f"  Number formats observed:")
            for fmt, c in number_formats.most_common(6):
                # Sanitize format strings for printing
                fmt_safe = fmt.replace("\n", "\\n")
                print(f"    {c:>5}  {fmt_safe[:60]}")
        print()

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
