"""
build_capability_matrix_xlsx.py
Creates 4 Excel spreadsheets from TES17 capability-matrix markdown files.
"""

import re
import os
import argparse
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
NAVY       = "1F3864"
LIGHT_BLUE = "DCE6F1"
WHITE      = "FFFFFF"
GREEN      = "C6EFCE"
YELLOW     = "FFEB9C"
RED        = "FFC7CE"
ORANGE     = "FFEB9C"   # reuse yellow for LOW-MEDIUM

HEADER_FONT   = Font(name="Arial", bold=True, color=WHITE, size=10)
BODY_FONT     = Font(name="Arial", size=10)
HEADER_FILL   = PatternFill("solid", fgColor=NAVY)
ALT_FILL      = PatternFill("solid", fgColor=LIGHT_BLUE)
WHITE_FILL    = PatternFill("solid", fgColor=WHITE)
GREEN_FILL    = PatternFill("solid", fgColor=GREEN)
YELLOW_FILL   = PatternFill("solid", fgColor=YELLOW)
RED_FILL      = PatternFill("solid", fgColor=RED)

WRAP  = Alignment(wrap_text=True, vertical="top")
CENTER= Alignment(wrap_text=True, vertical="center", horizontal="center")

# ---------------------------------------------------------------------------
# Helper: thin border
# ---------------------------------------------------------------------------
def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

# ---------------------------------------------------------------------------
# Coverage / gap colour for a cell value
# ---------------------------------------------------------------------------
def coverage_fill(value: str):
    v = value.strip().upper().strip("*")
    if v in ("STRONG", "HIGH"):
        return GREEN_FILL
    if v in ("PARTIAL", "MEDIUM", "LOW-MEDIUM"):
        return YELLOW_FILL
    if v in ("GAP", "LOW", "CRITICAL"):
        return RED_FILL
    # bolded GAP markers like **GAP**
    if "GAP" in v:
        return RED_FILL
    return None

# ---------------------------------------------------------------------------
# Apply header row
# ---------------------------------------------------------------------------
def apply_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font    = HEADER_FONT
        cell.fill    = HEADER_FILL
        cell.alignment = CENTER
        cell.border  = thin_border()

# ---------------------------------------------------------------------------
# Apply body row
# ---------------------------------------------------------------------------
def apply_body_row(ws, row_idx, values, alt=False, coverage_cols=None):
    fill = ALT_FILL if alt else WHITE_FILL
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font      = BODY_FONT
        cell.alignment = WRAP
        cell.border    = thin_border()
        # coverage colouring overrides alternating
        if coverage_cols and col in coverage_cols:
            cf = coverage_fill(val or "")
            cell.fill = cf if cf else fill
        else:
            cell.fill = fill

# ---------------------------------------------------------------------------
# Auto-size columns with caps
# ---------------------------------------------------------------------------
COL_CAPS = {
    "SOO Requirement": 42,
    "Requirement Type": 18,
    "Company Capability": 38,
    "Coverage": 14,
    "Gap?": 38,
    "SOO Outcome": 38,
    "Coverage Level": 16,
    "Primary Company Capability": 38,
    "Gap / Team Coverage Needed": 38,
    "Gap": 28,
    "Description": 40,
    "Severity": 14,
    "Recommended Teaming Partner": 26,
    "Partner Capability": 35,
    "Component": 30,
    "Current TRL": 14,
    "5-Month Sprint Target TRL": 20,
    "Basis": 38,
    "Differentiator": 28,
    "Company Advantage": 40,
    "Risk": 38,
}
DEFAULT_CAP = 35

def auto_size(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        header_val = col_cells[0].value or ""
        cap = COL_CAPS.get(str(header_val), DEFAULT_CAP)
        for cell in col_cells:
            if cell.value:
                for line in str(cell.value).split("\n"):
                    max_len = max(max_len, len(line))
        # clamp
        width = min(max(max_len + 2, 10), cap)
        ws.column_dimensions[col_letter].width = width

# ---------------------------------------------------------------------------
# Parse markdown table → list of dicts
# ---------------------------------------------------------------------------
def parse_md_table(lines):
    """Given a list of raw markdown lines for one table, return header + rows."""
    table_lines = [l for l in lines if l.strip().startswith("|")]
    if not table_lines:
        return [], []
    # header
    header = [c.strip() for c in table_lines[0].strip("|").split("|")]
    # skip separator row
    rows = []
    for line in table_lines[2:]:
        cells = [c.strip() for c in line.strip("|").split("|")]
        # pad to header length
        while len(cells) < len(header):
            cells.append("")
        rows.append(cells[:len(header)])
    return header, rows

# ---------------------------------------------------------------------------
# Split markdown into named sections
# ---------------------------------------------------------------------------
def split_sections(text):
    """Returns dict: section_num -> (title, lines[])"""
    sections = {}
    current_num = None
    current_title = ""
    current_lines = []

    section_re = re.compile(r"^##\s+Section\s+(\d+):\s*(.*)", re.IGNORECASE)

    for line in text.splitlines():
        m = section_re.match(line)
        if m:
            if current_num is not None:
                sections[int(current_num)] = (current_title, current_lines)
            current_num = m.group(1)
            current_title = m.group(2).strip()
            current_lines = []
        elif current_num is not None:
            current_lines.append(line)

    if current_num is not None:
        sections[int(current_num)] = (current_title, current_lines)

    return sections

# ---------------------------------------------------------------------------
# Parse numbered list (Section 4 Win Themes)
# ---------------------------------------------------------------------------
def parse_numbered_list(lines):
    items = []
    for line in lines:
        m = re.match(r"^\s*\d+\.\s+(.*)", line)
        if m:
            items.append(m.group(1).strip())
    return items

# ---------------------------------------------------------------------------
# Extract overall coverage % from section 1 text
# ---------------------------------------------------------------------------
def extract_coverage_pct(lines):
    for line in lines:
        m = re.search(r"Overall Coverage:\s*(.*?)%?\s*native", line, re.IGNORECASE)
        if m:
            return m.group(0).strip()
    return "See Section 1"

# ---------------------------------------------------------------------------
# Create a sheet from a table
# ---------------------------------------------------------------------------
def add_table_sheet(wb, sheet_name, headers, rows, coverage_cols=None):
    ws = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = "A2"
    apply_header(ws, headers)
    for i, row in enumerate(rows):
        apply_body_row(ws, i + 2, row, alt=(i % 2 == 1), coverage_cols=coverage_cols)
    ws.row_dimensions[1].height = 30
    auto_size(ws)
    return ws

# ---------------------------------------------------------------------------
# Create Win Themes sheet (numbered list)
# ---------------------------------------------------------------------------
def add_win_themes_sheet(wb, sheet_name, items):
    ws = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = "A2"
    apply_header(ws, ["#", "Win Theme"])
    for i, item in enumerate(items):
        apply_body_row(ws, i + 2, [str(i + 1), item], alt=(i % 2 == 1))
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 70
    ws.row_dimensions[1].height = 25
    return ws

# ---------------------------------------------------------------------------
# Create Summary sheet
# ---------------------------------------------------------------------------
def add_summary_sheet(wb, proposal_info, sections):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.freeze_panes = "A2"

    # Title row
    title_cell = ws.cell(row=1, column=1, value="CAPABILITY MATRIX SUMMARY")
    title_cell.font = Font(name="Arial", bold=True, color=WHITE, size=13)
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 35

    # Key fields
    fields = [
        ("Proposal", proposal_info["slug"]),
        ("Title", proposal_info["title"]),
        ("SOO Reference", proposal_info["soo_ref"]),
        ("Generated", proposal_info["generated"]),
        ("Overall Coverage", proposal_info["coverage"]),
        ("Priority Ranking", proposal_info.get("priority", "See Section 6")),
    ]

    row = 2
    for i, (label, value) in enumerate(fields):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font  = Font(name="Arial", bold=True, size=10)
        lc.fill  = ALT_FILL if i % 2 == 0 else WHITE_FILL
        lc.border = thin_border()
        lc.alignment = WRAP

        vc = ws.cell(row=row, column=2, value=value)
        vc.font  = BODY_FONT
        vc.fill  = ALT_FILL if i % 2 == 0 else WHITE_FILL
        vc.border = thin_border()
        vc.alignment = WRAP
        ws.merge_cells(f"B{row}:D{row}")
        row += 1

    # Gaps summary header
    row += 1
    gh = ws.cell(row=row, column=1, value="Key Gaps (see Sheet 3 for details)")
    gh.font  = Font(name="Arial", bold=True, color=WHITE, size=10)
    gh.fill  = HEADER_FILL
    gh.alignment = CENTER
    ws.merge_cells(f"A{row}:D{row}")
    ws.row_dimensions[row].height = 25
    row += 1

    # Pull gap rows from section 3
    sec3_title, sec3_lines = sections.get(3, ("", []))
    h3, r3 = parse_md_table(sec3_lines)
    if r3:
        # condensed: Gap | Severity | Partner
        gap_i = next((i for i, c in enumerate(h3) if "Gap" == c), 0)
        sev_i = next((i for i, c in enumerate(h3) if "Severity" in c), 3)
        ptn_i = next((i for i, c in enumerate(h3) if "Recommended" in c), 4)

        gh2 = ws.cell(row=row, column=1, value="Gap")
        gh2.font = HEADER_FONT; gh2.fill = HEADER_FILL; gh2.border = thin_border(); gh2.alignment = CENTER
        gs2 = ws.cell(row=row, column=2, value="Severity")
        gs2.font = HEADER_FONT; gs2.fill = HEADER_FILL; gs2.border = thin_border(); gs2.alignment = CENTER
        gp2 = ws.cell(row=row, column=3, value="Teaming Partner")
        gp2.font = HEADER_FONT; gp2.fill = HEADER_FILL; gp2.border = thin_border(); gp2.alignment = CENTER
        ws.merge_cells(f"C{row}:D{row}")
        row += 1

        for gi, gr in enumerate(r3):
            sev_val = gr[sev_i] if sev_i < len(gr) else ""
            sev_fill = coverage_fill(sev_val) or (ALT_FILL if gi % 2 == 1 else WHITE_FILL)
            gc = ws.cell(row=row, column=1, value=gr[gap_i] if gap_i < len(gr) else "")
            gc.font = BODY_FONT; gc.fill = ALT_FILL if gi%2==1 else WHITE_FILL; gc.border=thin_border(); gc.alignment=WRAP
            sc = ws.cell(row=row, column=2, value=sev_val.strip("*"))
            sc.font = BODY_FONT; sc.fill = sev_fill; sc.border=thin_border(); sc.alignment=WRAP
            pc = ws.cell(row=row, column=3, value=gr[ptn_i] if ptn_i < len(gr) else "")
            pc.font = BODY_FONT; pc.fill = ALT_FILL if gi%2==1 else WHITE_FILL; pc.border=thin_border(); pc.alignment=WRAP
            ws.merge_cells(f"C{row}:D{row}")
            row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 12
    return ws

# ---------------------------------------------------------------------------
# Extract proposal metadata from header block
# ---------------------------------------------------------------------------
def extract_metadata(text):
    slug = re.search(r"\*\*Proposal:\*\*\s*(\S+)", text)
    generated = re.search(r"\*\*Generated:\*\*\s*(\S+)", text)
    soo_ref = re.search(r"\*\*SOO Reference:\*\*\s*(.*)", text)
    title_m = re.search(r"^#\s+Capability Matrix\s+[—–-]+\s*(.*)", text, re.MULTILINE)
    coverage_m = re.search(r"Overall Coverage:\s*~?([\d]+%?\s*native[^|*\n]*)", text)
    priority_m = re.search(r"Priority ranking[^\n]*\n.*?1\.\s*(tes17[^\s(]+)", text, re.DOTALL | re.IGNORECASE)
    return {
        "slug": slug.group(1) if slug else "",
        "generated": generated.group(1) if generated else "2026-05-13",
        "soo_ref": soo_ref.group(1).strip() if soo_ref else "",
        "title": title_m.group(1).strip() if title_m else "",
        "coverage": ("~" + coverage_m.group(1).strip()) if coverage_m else "See Section 1",
        "priority": "#1 of 4" if priority_m and priority_m.group(1) in (slug.group(1) if slug else "") else "See Section 6/7",
    }

# ---------------------------------------------------------------------------
# Main build for one file
# ---------------------------------------------------------------------------
def build_xlsx(md_path: Path, out_path: Path):
    text = md_path.read_text(encoding="utf-8")
    sections = split_sections(text)
    meta = extract_metadata(text)

    # Determine priority from bid decision section (mesh-intel only, Section 7)
    if "tes17-aa-cognitive-partner" in str(md_path):
        meta["priority"] = "#1 of 4 (highest alignment)"
    elif "tes17-edge-federated" in str(md_path):
        meta["priority"] = "#2 of 4"
    elif "tes17-edge-data-curation" in str(md_path):
        meta["priority"] = "#3 of 4"
    elif "tes17-aa-mesh-intel" in str(md_path):
        meta["priority"] = "#4 of 4 (most teaming-dependent)"

    # Pull coverage % from section 1 note line
    s1_title, s1_lines = sections.get(1, ("", []))
    meta["coverage"] = extract_coverage_pct(s1_lines) or meta["coverage"]

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ------- Summary (index 0) -------
    add_summary_sheet(wb, meta, sections)

    # ------- Sheet 1: SOO Mapping -------
    s1_title, s1_lines = sections.get(1, ("", []))
    h1, r1 = parse_md_table(s1_lines)
    if h1:
        cov_col_1 = {i+1 for i, c in enumerate(h1) if c in ("Coverage",)}
        add_table_sheet(wb, "1-SOO Mapping", h1, r1, coverage_cols=cov_col_1)

    # ------- Sheet 2: Outcome Coverage -------
    s2_title, s2_lines = sections.get(2, ("", []))
    h2, r2 = parse_md_table(s2_lines)
    if h2:
        cov_col_2 = {i+1 for i, c in enumerate(h2) if "Coverage" in c}
        add_table_sheet(wb, "2-Outcome Coverage", h2, r2, coverage_cols=cov_col_2)

    # ------- Sheet 3: Gaps & Teaming -------
    s3_title, s3_lines = sections.get(3, ("", []))
    h3, r3 = parse_md_table(s3_lines)
    if h3:
        cov_col_3 = {i+1 for i, c in enumerate(h3) if "Severity" in c}
        add_table_sheet(wb, "3-Gaps & Teaming", h3, r3, coverage_cols=cov_col_3)

    # ------- Sheet 4: Win Themes -------
    s4_title, s4_lines = sections.get(4, ("", []))
    items4 = parse_numbered_list(s4_lines)
    add_win_themes_sheet(wb, "4-Win Themes", items4)

    # ------- Sheet 5: TRL Assessment -------
    s5_title, s5_lines = sections.get(5, ("", []))
    h5, r5 = parse_md_table(s5_lines)
    if h5:
        add_table_sheet(wb, "5-TRL Assessment", h5, r5)

    # ------- Sheet 6: Competitive Diff -------
    s6_title, s6_lines = sections.get(6, ("", []))
    h6, r6 = parse_md_table(s6_lines)
    if h6:
        add_table_sheet(wb, "6-Competitive Diff", h6, r6)

    # ------- Sheet 7: Open Questions (or bid decision notes) -------
    # mesh-intel has Section 7 = Bid Decision, Section 8 = Open Questions
    is_mesh = "tes17-aa-mesh-intel" in str(md_path)
    if is_mesh:
        s7_title, s7_lines = sections.get(7, ("", []))
        s8_title, s8_lines = sections.get(8, ("", []))
        # Section 7 bid decision is prose — put it as a notes sheet
        ws7 = wb.create_sheet(title="7-Bid Decision")
        ws7.freeze_panes = "A2"
        apply_header(ws7, ["Bid Decision Notes"])
        row_idx = 2
        for line in s7_lines:
            line = line.strip()
            if line and not line.startswith("---"):
                c = ws7.cell(row=row_idx, column=1, value=line)
                c.font = BODY_FONT; c.fill = WHITE_FILL if row_idx % 2 == 0 else ALT_FILL
                c.alignment = WRAP; c.border = thin_border()
                row_idx += 1
        ws7.column_dimensions["A"].width = 90
        ws7.row_dimensions[1].height = 25

        # Section 8 open questions
        items8 = parse_numbered_list(s8_lines)
        ws8 = wb.create_sheet(title="8-Open Questions")
        ws8.freeze_panes = "A2"
        apply_header(ws8, ["#", "Open Question / Assumption"])
        for i, item in enumerate(items8):
            apply_body_row(ws8, i + 2, [str(i + 1), item], alt=(i % 2 == 1))
        ws8.column_dimensions["A"].width = 5
        ws8.column_dimensions["B"].width = 80
        ws8.row_dimensions[1].height = 25
    else:
        # Standard: Section 7 = Open Questions numbered list
        s7_title, s7_lines = sections.get(7, ("", []))
        items7 = parse_numbered_list(s7_lines)
        ws7 = wb.create_sheet(title="7-Open Questions")
        ws7.freeze_panes = "A2"
        apply_header(ws7, ["#", "Open Question / Assumption"])
        for i, item in enumerate(items7):
            apply_body_row(ws7, i + 2, [str(i + 1), item], alt=(i % 2 == 1))
        ws7.column_dimensions["A"].width = 5
        ws7.column_dimensions["B"].width = 80
        ws7.row_dimensions[1].height = 25

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"  Saved: {out_path}")

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Build a formatted capability-matrix workbook from markdown."
    )
    parser.add_argument("input", type=Path, help="Path to working/capability-matrix.md")
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        help="Output .xlsx path. Defaults to final/xlsx/capability-matrix.xlsx next to the proposal.",
    )
    args = parser.parse_args()

    md_path = args.input
    proposal_root = md_path.parent.parent if md_path.parent.name == "working" else md_path.parent
    out_path = args.output or proposal_root / "final" / "xlsx" / "capability-matrix.xlsx"

    print(f"\nBuilding: {proposal_root.name}")
    build_xlsx(md_path, out_path)
    print("\nAll done.")
