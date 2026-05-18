"""Build milestone-schedule.xlsx for digital-guardian-sim OTA proposal."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

out_path = Path("proposals/digital-guardian-sim/final/xlsx/milestone-schedule.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)

wb = openpyxl.Workbook()

header_fill = PatternFill("solid", fgColor="1F3864")
accent_fill = PatternFill("solid", fgColor="2E75B6")
alt_fill    = PatternFill("solid", fgColor="D6E4F0")
total_fill  = PatternFill("solid", fgColor="1F3864")
white_font  = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="AAAAAA")
bord = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = white_font
    c.fill = accent_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = bord
    return c


def dat(ws, row, col, value, align="left", bold=False, fill=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
    c.border = bord
    if bold:
        c.font = Font(bold=True)
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    return c


# ── Sheet 1: Milestone Summary ───────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Milestone Summary"

ws1.merge_cells("A1:H1")
t = ws1["A1"]
t.value = "Digital Guardian Simulator MVP — Milestone Payment Schedule"
t.font = Font(size=14, bold=True, color="FFFFFF")
t.fill = header_fill
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

ws1.merge_cells("A2:H2")
s = ws1["A2"]
s.value = "[Your Company], Inc.  |  Total: $[total amount]  |  Period of Performance: [N] Weeks  |  OTA via [vehicle]"
s.font = Font(italic=True, color="FFFFFF")
s.fill = accent_fill
s.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 18

for i, h in enumerate(["Milestone", "Week", "Title", "Direct Labor", "ODC", "Subtotal", "Fee", "Payment"], 1):
    hdr(ws1, 3, i, h)
ws1.row_dimensions[3].height = 18

milestones = [
    ("M1", "Week 2",  "Environment + Opening Statement Demo",           21150, 1500,  22650, 2350,  25000),
    ("M2", "Week 5",  "Core Orchestration + Why State v1 + Go/No-Go",   45700, 1800,  47500, 7500,  55000),
    ("M3", "Week 7",  "Full Injection + Real-Time Why + 45-Min Test",    44650, 350,   45000, 10000, 55000),
    ("M4", "Week 8",  "Delivery Package + Documentation + Demo Support", 15450, 4050,  19500, 10500, 30000),
]

for r, (ms, wk, title, dl, odc, sub, fee, pmt) in enumerate(milestones, 4):
    dat(ws1, r, 1, ms,    align="center", bold=True)
    dat(ws1, r, 2, wk,    align="center")
    dat(ws1, r, 3, title)
    for ci, v in enumerate([dl, odc, sub, fee, pmt], 4):
        c = ws1.cell(row=r, column=ci, value=v)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="top")
        c.border = bord
        if ci == 8:
            c.font = Font(bold=True)

tr = len(milestones) + 4
ws1.merge_cells(f"A{tr}:C{tr}")
tc = ws1[f"A{tr}"]
tc.value = "TOTAL"
tc.font = white_font
tc.fill = total_fill
tc.alignment = Alignment(horizontal="center", vertical="center")
tc.border = bord
ws1[f"B{tr}"].border = bord
ws1[f"C{tr}"].border = bord
for ci, v in enumerate([127000, 7700, 134650, 30350, 165000], 4):
    c = ws1.cell(row=tr, column=ci, value=v)
    c.number_format = '"$"#,##0'
    c.font = white_font
    c.fill = total_fill
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = bord
ws1.row_dimensions[tr].height = 18

for i, w in enumerate([8, 10, 42, 14, 12, 14, 12, 14], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 2: Acceptance Criteria ────────────────────────────────────────────
ws2 = wb.create_sheet("Acceptance Criteria")

ws2.merge_cells("A1:E1")
t2 = ws2["A1"]
t2.value = "Milestone Acceptance Criteria"
t2.font = Font(size=13, bold=True, color="FFFFFF")
t2.fill = header_fill
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

for i, h in enumerate(["MS", "Week", "Key Deliverables", "Acceptance Criterion", "GFI Dependency"], 1):
    hdr(ws2, 2, i, h)

criteria = [
    ("M1", "Wk 2",
     "Docker Compose baseline; vLLM inference server on DGX Spark; persona Opening Round v1",
     "Four AI personas produce role-appropriate Opening Round statements without critical failure. Government IT administrator can bring up the stack with a single docker-compose up command.",
     "DGX Spark provisioned within 3 business days of contract award"),
    ("M2", "Wk 5",
     "Guardian Engine v1 (state machine); Why State Manager v1 (READ); Visualization Dashboard v1; Go/No-Go decision documented",
     "Complete Status Round session with all four personas and Why visualization rendering. Government and the company jointly execute Go/No-Go: proceed with full Why State v2 or activate descope path.",
     "DAU WAS curriculum and DAF policy corpus delivered by Day 1"),
    ("M3", "Wk 7",
     "Why State Manager v2 (WRITE + WebSocket); Injection API; ZT audit log; 45-min full session test",
     "Three complete 45-minute war game sessions with minimum 4 audience-driven injections each. Why priority bars reorder within <500ms. Audit log captures all model I/O. Zero external API calls during execution.",
     "None (all GFI consumed by M2)"),
    ("M4", "Wk 8",
     "Final Docker Compose package; ops runbook (min 20 pages); on-site or remote demo support",
     "Government IT admin with Docker familiarity deploys and operates without company assistance. One company technical representative available for the demonstration.",
     "Demo date confirmed by government by Week 6"),
]

for r, row in enumerate(criteria, 3):
    for c, val in enumerate(row, 1):
        cell_obj = ws2.cell(row=r, column=c, value=val)
        cell_obj.alignment = Alignment(
            horizontal="center" if c <= 2 else "left", vertical="top", wrap_text=True)
        cell_obj.border = bord
        if c == 1:
            cell_obj.font = Font(bold=True)
    ws2.row_dimensions[r].height = 80

for i, w in enumerate([6, 8, 40, 54, 32], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 3: Labor Buildup ───────────────────────────────────────────────────
ws3 = wb.create_sheet("Labor Buildup")

ws3.merge_cells("A1:F1")
t3 = ws3["A1"]
t3.value = "Labor Hour Buildup by Milestone"
t3.font = Font(size=13, bold=True, color="FFFFFF")
t3.fill = header_fill
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 26

for i, h in enumerate(["Labor Category", "Loaded Rate", "M1 Hours", "M2 Hours", "M3 Hours", "M4 Hours"], 1):
    hdr(ws3, 2, i, h)

labor = [
    ("ML/AI Engineering Lead",  200, 40,  100, 90,  20),
    ("ML Engineer",              175, 40,   80, 70,  20),
    ("Full-stack / DevOps Eng.", 165, 30,   60, 80,  30),
    ("PM / Technical Writer",    150,  8,   12,  8,  20),
]

for r, (cat, rate, m1, m2, m3, m4) in enumerate(labor, 3):
    dat(ws3, r, 1, cat)
    c2 = ws3.cell(row=r, column=2, value=rate)
    c2.number_format = '"$"#,##0"/hr"'
    c2.alignment = Alignment(horizontal="center")
    c2.border = bord
    for ci, v in enumerate([m1, m2, m3, m4], 3):
        c = ws3.cell(row=r, column=ci, value=v)
        c.alignment = Alignment(horizontal="center")
        c.border = bord

tr3 = len(labor) + 3
dat(ws3, tr3, 1, "TOTAL HOURS", bold=True)
ws3.cell(row=tr3, column=2, value="760 total").border = bord
ws3.cell(row=tr3, column=2).alignment = Alignment(horizontal="center")
for ci, v in enumerate([118, 252, 248, 90], 3):
    c = ws3.cell(row=tr3, column=ci, value=v)
    c.font = Font(bold=True)
    c.fill = alt_fill
    c.alignment = Alignment(horizontal="center")
    c.border = bord

for i, w in enumerate([28, 14, 12, 12, 12, 12], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w


wb.save(out_path)
print(f"[OK] {out_path}")
print(f"     Sheets: {[ws.title for ws in wb.worksheets]}")
